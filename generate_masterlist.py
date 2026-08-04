#!/usr/bin/env python3
"""
generate_masterlist.py
======================
Konvertiert Inventarlisten in masterlist.json für die Inventur-PWA.

Primärquelle: wöchentliche Inventarliste aus PART_MGMT/weekly_inventorylist/ (per Mail von
giovanni.fiore@canon.ch bzw. ch-sphd@canon.ch zugestellt, siehe
inventar_mail_watch.py). Das Spaltenlayout wird über Namens-Aliase gesucht statt
über feste Indizes, weil diese Datei nicht von uns erzeugt wird und ihr Format
über die Wochen schon mehrfach leicht variiert hat (Spaltenreihenfolge,
Spaltennamen, teils fehlende Spalten).

Fallback: "Inventory value by item number, planned status, and age.xlsx",
liegt seit 2026-08 im SELBEN Ordner (PART_MGMT/weekly_inventorylist/) wie die wöchentliche
Liste, wird aber nie als wöchentlicher Snapshot missverstanden (fester
Dateiname, siehe find_latest_weekly_file). Kommt nur zum Zug, wenn die
wöchentliche Liste fehlt oder ihr Format nicht zu den erwarteten
Pflichtspalten passt - UND diese Fallback-Datei (per Datei-Änderungsdatum)
neuer ist als der bereits vorhandene masterlist.json-Stand. Ist sie das nicht
(z.B. eine alte, liegengebliebene Kopie), passiert nichts - masterlist.json
bleibt unverändert, kein Rückschritt zu älteren Daten. In jedem Fehlerfall
wird eine Alarm-Datei (_masterlist_alarm.txt in PART_MGMT/weekly_inventorylist/)
geschrieben, die bestehen bleibt bis die wöchentliche Liste wieder
erfolgreich gelesen wird.

Aufruf:
    python generate_masterlist.py
    python generate_masterlist.py --input "pfad/zur/alten/fallback-datei.xlsx"
    python generate_masterlist.py --inventar-dir "pfad/zum/inventar/ordner"
    python generate_masterlist.py --watch

Wöchentlich automatisch ausgeführt (via inventar_mail_watch.py + --watch).
"""

import sys
import os
import re
import json
import argparse
import subprocess
import datetime
import time
import socket
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Fehlende Abhängigkeit: openpyxl")
    print("  → pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Standardpfade (relativ zu diesem Script)
# ---------------------------------------------------------------------------
PART_MGMT_DIR    = Path(__file__).parent.parent / "PART_MGMT"
DEFAULT_INVENTAR = PART_MGMT_DIR / "weekly_inventorylist"
OLD_SOURCE_NAME  = "Inventory value by item number, planned status, and age.xlsx"
DEFAULT_INPUT    = DEFAULT_INVENTAR / OLD_SOURCE_NAME
DEFAULT_OUT      = Path(__file__).parent / "data" / "masterlist.json"
ALARM_FILENAME   = "_masterlist_alarm.txt"

# ---------------------------------------------------------------------------
# Spaltenindizes für die ALTE Quelle (0-basiert, festes Layout, nur Fallback)
# ---------------------------------------------------------------------------
COL_SUB   = 3   # SUB INVENTORY           z.B. "CHU.8678"
COL_ENG   = 4   # ENGINEER/LOCATION NAME  z.B. "8678 Edubook CSP341555xxx"
COL_ITEM  = 7   # ITEM NUMBER             z.B. "0001070128700"
COL_NAME  = 8   # ITEM NAME               z.B. "BELT-TIMING-1830-2MR09"
COL_MIN   = 17  # MIN QTY
COL_QOH   = 19  # TOTAL ON HAND QTY
COL_VAL   = 21  # TOTAL STOCK VALUE
COL_AGE   = 23  # AGE RANK                z.B. "< 30 DAYS", "Excluded"

# ---------------------------------------------------------------------------
# Spalten-Aliase für die NEUE wöchentliche Quelle (per Name gesucht, da
# Reihenfolge und leichte Namensvarianten über die Wochen schon vorkamen).
# 'max' und 'age'/'eng' sind optional - fehlen sie, wird einfach ohne sie
# weitergemacht statt abzubrechen.
# ---------------------------------------------------------------------------
WEEKLY_COL_ALIASES = {
    'sub':  ['SUB INVENTORY', 'Sub'],
    'item': ['ITEM NUMBER', 'Item'],
    'name': ['ITEM NAME', 'Item Description'],
    'min':  ['MIN QTY'],
    'max':  ['MAX QTY'],
    'qoh':  ['TOTAL ON HAND QTY', 'TOTAL ON HAND QTY FAMILY', 'On-hand'],
    'age':  ['AGE RANK'],
    'eng':  ['ENGINEER/LOCATION NAME'],
}
WEEKLY_REQUIRED_COLS = ['sub', 'item', 'name', 'min', 'qoh']


class WeeklyFormatError(Exception):
    """Wöchentliche Liste fehlt oder hat ein unerwartetes Spaltenlayout.
    Löst den Fallback auf die alte Quelle + einen Alarm aus."""


def parse_age_rank(raw) -> int:
    """
    Konvertiert AGE RANK String zu int:
      -1 = Excluded / unbekannt
       0 = < 30 days
       1 = 30 - 60 days
       2 = 61 - 90 days
       3 = 91 - 180 days
       4 = 181 - 360 days
       5 = > 360 days
    """
    if raw is None:
        return -1
    s = str(raw).strip().upper()
    # Wöchentliche Liste stellt eine Rang-Nummer voran, z.B. "1. <30 DAYS"
    s = re.sub(r'^\d+\.\s*', '', s)
    if not s or 'EXCLUDED' in s:
        return -1
    if '< 30' in s or s.startswith('<30'):
        return 0
    if '30' in s and '60' in s:
        return 1
    if '61' in s and '90' in s:
        return 2
    if '91' in s and '180' in s:
        return 3
    if '181' in s and '360' in s:
        return 4
    if '> 360' in s or '>360' in s:
        return 5
    return -1


def clean_location_name(raw: str) -> str:
    """
    Kürzt den Engineer/Location-Namen auf etwas Lesbares.
    "8678 Edubook CSP341555xxx"  →  "Edubook"
    "Baudat"                     →  "Baudat"
    """
    name = raw.split('CSP')[0].strip()
    name = re.sub(r'^\d{4,5}\s+', '', name).strip()
    return name or raw


def has_internet(host='github.com', port=443, timeout=5) -> bool:
    try:
        socket.setdefaulttimeout(timeout)
        socket.socket(socket.AF_INET, socket.SOCK_STREAM).connect((host, port))
        return True
    except OSError:
        return False


def normalize_item_number(raw) -> str:
    """Führende Nullen entfernen; alphanumerische Nummern bleiben unverändert."""
    s = str(raw).strip() if raw is not None else ''
    return s.lstrip('0') or s


def parse_snapshot_date(filename: str):
    """
    Extrahiert das Snapshot-Datum aus wechselnden Dateinamen-Formaten, z.B.:
      'Inventory 28072026.xlsx'        -> DDMMYYYY
      'Inventarliste210726.xlsx'       -> DDMMYY
      'Inventorylist 20.01.2026.xlsx'  -> DD.MM.YYYY
    Gibt None zurück wenn kein Datum erkannt wird.
    """
    m = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', filename)
    if m:
        d, mo, y = m.groups()
        try:
            return datetime.date(int(y), int(mo), int(d))
        except ValueError:
            pass
    m = re.search(r'(\d{2})(\d{2})(\d{4})', filename)
    if m:
        d, mo, y = m.groups()
        try:
            return datetime.date(int(y), int(mo), int(d))
        except ValueError:
            pass
    m = re.search(r'(\d{2})(\d{2})(\d{2})(?!\d)', filename)
    if m:
        d, mo, y = m.groups()
        try:
            return datetime.date(2000 + int(y), int(mo), int(d))
        except ValueError:
            pass
    return None


def find_latest_weekly_file(inventar_dir: Path):
    """
    Neueste Inventarliste im inventar-Ordner (per Datum im Dateinamen).
    Ignoriert Dateien ohne 'invent' im Namen (z.B. Auswertungen wie
    Nullbestand_Report_*.xlsx), die alte Fallback-Quelle (fester Dateiname,
    liegt seit 2026-08 im selben Ordner) und Dateien deren Datum sich nicht
    parsen lässt. Gibt still None zurück wenn nichts gefunden wird -
    Logging/Fehlermeldungen übernimmt der Aufrufer (convert_weekly).
    """
    if not inventar_dir.exists():
        return None
    candidates = []
    for f in inventar_dir.glob('*.xlsx'):
        if f.name.startswith('~$'):  # Office-Lock-Dateien
            continue
        if f.name == OLD_SOURCE_NAME:  # alte Fallback-Quelle, nie als weekly werten
            continue
        if 'invent' not in f.name.lower():
            continue
        d = parse_snapshot_date(f.name)
        if d is None:
            continue
        candidates.append((d, f))
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0])
    return candidates[-1][1]


def _find_col_indices(header, aliases_map):
    """Ordnet kanonische Feldnamen ihrem Spaltenindex zu (per Name gesucht)."""
    header_norm = [str(h).strip() if h is not None else '' for h in header]
    idx = {}
    for canon, aliases in aliases_map.items():
        for a in aliases:
            if a in header_norm:
                idx[canon] = header_norm.index(a)
                break
    return idx


def convert_weekly(inventar_dir: Path):
    """
    Liest die neueste wöchentliche Inventarliste aus inventar_dir.
    Wirft WeeklyFormatError wenn keine Datei gefunden wird, die Datei leer
    ist, oder eine Pflichtspalte fehlt.
    """
    path = find_latest_weekly_file(inventar_dir)
    if path is None:
        raise WeeklyFormatError(
            f"Keine wöchentliche Inventarliste in {inventar_dir} gefunden "
            f"(erwartet: *.xlsx mit 'invent' im Namen und erkennbarem Datum)."
        )

    print(f"Lese (wöchentliche Quelle): {path.name} ...", flush=True)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        wb.close()
        raise WeeklyFormatError(f"{path.name} ist leer (keine Kopfzeile).")

    idx = _find_col_indices(header, WEEKLY_COL_ALIASES)
    missing = [c for c in WEEKLY_REQUIRED_COLS if c not in idx]
    if missing:
        wb.close()
        raise WeeklyFormatError(
            f"{path.name}: Pflichtspalten fehlen: {missing}. "
            f"Gefundene Spalten: {list(header)}"
        )

    stand_date = parse_snapshot_date(path.name)
    stand_str = stand_date.isoformat() if stand_date else \
        datetime.datetime.fromtimestamp(os.path.getmtime(str(path))).date().isoformat()

    item_map   = defaultdict(dict)   # item_nr → {sub: {min, qoh, [a]}}
    item_names = {}                  # item_nr → name
    eng_names  = {}                  # sub → short location name
    row_count  = 0

    for row in rows:
        if row is None:
            continue

        sub_raw = row[idx['sub']] if idx['sub'] < len(row) else None
        sub = str(sub_raw).strip() if sub_raw else ''
        if not sub or sub == 'None':
            continue

        item_raw = row[idx['item']] if idx['item'] < len(row) else None
        if item_raw is None:
            continue
        item = normalize_item_number(item_raw)
        if not item or item == 'None':
            continue

        name_raw = row[idx['name']] if idx['name'] < len(row) else None
        name = str(name_raw).strip() if name_raw else ''

        mn_raw  = row[idx['min']] if idx['min'] < len(row) else None
        qoh_raw = row[idx['qoh']] if idx['qoh'] < len(row) else None
        mn  = int(mn_raw)  if isinstance(mn_raw, (int, float)) else 0
        qoh = int(qoh_raw) if isinstance(qoh_raw, (int, float)) else 0

        entry = {'min': mn, 'qoh': qoh}
        if 'age' in idx and idx['age'] < len(row):
            entry['a'] = parse_age_rank(row[idx['age']])
        # 'v' (Preis) bewusst weggelassen - die wöchentliche Liste hat keine
        # Preisspalte. index.html prüft `loc.v != null` und fällt dann auf
        # '—' zurück, MIN/QOH/Fehlend-Logik bleiben unberührt.

        item_map[item][sub] = entry

        if item not in item_names:
            item_names[item] = name

        if sub not in eng_names:
            if 'eng' in idx and idx['eng'] < len(row) and row[idx['eng']]:
                eng_names[sub] = clean_location_name(str(row[idx['eng']]).strip())
            else:
                eng_names[sub] = sub

        row_count += 1

    wb.close()

    if row_count == 0:
        raise WeeklyFormatError(f"{path.name}: keine gültigen Datenzeilen gelesen.")

    print(f"  {row_count:,} Datenzeilen | {len(eng_names)} Lager | {len(item_map)} unique Artikel")

    data = {
        '_generated': datetime.date.today().isoformat(),
        '_stand':     stand_str,
        '_source':    f'weekly:{path.name}',
        '_w': eng_names,
        'i': {
            item: {'n': item_names[item], 'l': locs}
            for item, locs in item_map.items()
        }
    }
    return data


def convert_old(input_path: Path):
    """Alte Quelle (feste Spaltenindizes, ein Datenlieferant, ein Format) - Fallback."""
    print(f"Lese (alte Quelle, Fallback): {input_path.name} ...", flush=True)

    wb = openpyxl.load_workbook(str(input_path), read_only=True, data_only=True)
    ws = wb.active

    mtime = os.path.getmtime(str(input_path))
    stand_date = datetime.datetime.fromtimestamp(mtime).date().isoformat()
    print(f"  Stand (Quelldatei): {stand_date}")

    header = list(next(ws.iter_rows(values_only=True)))
    expected = {
        COL_SUB:  'SUB INVENTORY',
        COL_ITEM: 'ITEM NUMBER',
        COL_MIN:  'MIN QTY',
        COL_VAL:  'TOTAL STOCK VALUE',
        COL_AGE:  'AGE RANK',
    }
    for col, name in expected.items():
        actual = str(header[col]).strip() if col < len(header) else '?'
        if actual != name:
            print(f"⚠️  Spalte [{col}] erwartet '{name}', gefunden '{actual}'")
            print("   Bitte Spaltenindizes in generate_masterlist.py anpassen.")

    item_map   = defaultdict(dict)
    item_names = {}
    eng_names  = {}
    row_count  = 0

    for row in ws.iter_rows(values_only=True, min_row=2):
        if len(row) <= COL_QOH:
            continue

        sub = str(row[COL_SUB]).strip() if row[COL_SUB] else ''
        if not sub or sub == 'None':
            continue

        raw_item = str(row[COL_ITEM]).strip() if row[COL_ITEM] else ''
        item     = raw_item.lstrip('0') or raw_item
        if not item or item == 'None':
            continue

        name    = str(row[COL_NAME]).strip() if row[COL_NAME] else ''
        eng     = str(row[COL_ENG]).strip()  if row[COL_ENG]  else ''
        mn      = int(row[COL_MIN]) if isinstance(row[COL_MIN], (int, float)) else 0
        qoh     = int(row[COL_QOH]) if isinstance(row[COL_QOH], (int, float)) else 0
        val_raw = row[COL_VAL] if len(row) > COL_VAL else None
        age_raw = row[COL_AGE] if len(row) > COL_AGE else None
        val     = round(float(val_raw), 2) if isinstance(val_raw, (int, float)) else 0.0
        age     = parse_age_rank(age_raw)

        item_map[item][sub] = {'min': mn, 'qoh': qoh, 'a': age, 'v': val}

        if item not in item_names:
            item_names[item] = name
        if sub not in eng_names:
            eng_names[sub] = clean_location_name(eng)

        row_count += 1

    wb.close()

    print(f"  {row_count:,} Datenzeilen | {len(eng_names)} Lager | {len(item_map)} unique Artikel")

    data = {
        '_generated': datetime.date.today().isoformat(),
        '_stand':     stand_date,
        '_source':    f'fallback:{input_path.name}',
        '_w': eng_names,
        'i': {
            item: {'n': item_names[item], 'l': locs}
            for item, locs in item_map.items()
        }
    }
    return data


def read_existing_stand(out_path: Path):
    """Liest '_stand' aus einer bereits vorhandenen masterlist.json (falls
    vorhanden) - dient als Referenz um zu entscheiden ob eine Fallback-Datei
    ueberhaupt einen Fortschritt darstellt oder nur ein alter Stand waere."""
    if not out_path.exists():
        return None
    try:
        existing = json.loads(out_path.read_text(encoding='utf-8'))
        s = existing.get('_stand')
        if s:
            return datetime.date.fromisoformat(s)
    except (json.JSONDecodeError, OSError, ValueError):
        pass
    return None


def write_alarm(inventar_dir: Path, message: str, note: str = ''):
    inventar_dir.mkdir(parents=True, exist_ok=True)
    alarm_path = inventar_dir / ALARM_FILENAME
    ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    alarm_path.write_text(
        f"[{ts}] Wöchentliche Inventarliste konnte nicht gelesen werden:\n\n"
        f"{message}\n\n"
        f"{note}\n"
        f"Diese Datei bleibt bestehen bis das Problem behoben ist - sie wird "
        f"automatisch gelöscht, sobald die wöchentliche Liste wieder "
        f"erfolgreich gelesen wird.\n",
        encoding='utf-8'
    )
    print(f"🚨 ALARM geschrieben: {alarm_path}")


def clear_alarm(inventar_dir: Path):
    alarm_path = inventar_dir / ALARM_FILENAME
    if alarm_path.exists():
        alarm_path.unlink()
        print(f"✅ Alarm zurückgesetzt ({alarm_path.name} gelöscht) - wöchentliche Quelle funktioniert wieder.")


def watch_signature(input_path: Path, inventar_dir: Path):
    """Zustand beider Quellen für die Änderungserkennung im --watch-Modus."""
    sig = {}
    if input_path.exists():
        sig['old'] = os.path.getmtime(str(input_path))
    latest = find_latest_weekly_file(inventar_dir)
    if latest is not None:
        sig['weekly'] = (str(latest), os.path.getmtime(str(latest)))
    return sig


def run_watch(input_path: Path, inventar_dir: Path, out_path: Path, no_push: bool, interval: int):
    print(f"Überwache: {inventar_dir} (primär) und {input_path} (Fallback)")
    print(f"Prüfintervall: {interval}s  |  Ctrl+C zum Beenden.\n")

    last_sig = None

    try:
        while True:
            sig = watch_signature(input_path, inventar_dir)

            if last_sig is None:
                last_sig = sig  # Startzustand merken, nicht sofort konvertieren
            elif sig != last_sig:
                print(f"\n[{datetime.datetime.now():%H:%M:%S}] Dateiänderung erkannt.", flush=True)
                last_sig = sig

                if not no_push:
                    while not has_internet():
                        print(f"[{datetime.datetime.now():%H:%M:%S}] Kein Internet – warte 60s ...", flush=True)
                        time.sleep(60)

                convert(input_path, inventar_dir, out_path, no_push)

            time.sleep(interval)
    except KeyboardInterrupt:
        print("\nWatcher beendet.")


def convert(input_path: Path, inventar_dir: Path, out_path: Path, no_push: bool):
    try:
        data = convert_weekly(inventar_dir)
        clear_alarm(inventar_dir)
    except WeeklyFormatError as e:
        print(f"⚠️  Wöchentliche Quelle fehlgeschlagen: {e}")

        if not input_path.exists():
            write_alarm(inventar_dir, str(e),
                        f"Auch keine Fallback-Datei gefunden ({input_path.name}) - "
                        f"masterlist.json wurde NICHT veraendert.")
            print(f"❌ Auch Fallback-Datei nicht gefunden: {input_path} - nichts unternommen.")
            return

        fallback_date = datetime.datetime.fromtimestamp(os.path.getmtime(str(input_path))).date()
        existing_stand = read_existing_stand(out_path)
        if existing_stand is not None and fallback_date <= existing_stand:
            write_alarm(inventar_dir, str(e),
                        f"Fallback-Datei ({input_path.name}) ist vom {fallback_date}, "
                        f"bestehender masterlist.json-Stand ist bereits vom {existing_stand} "
                        f"- Fallback waere ein Rueckschritt, masterlist.json wurde NICHT veraendert.")
            print(f"   Fallback-Datei ({fallback_date}) ist nicht neuer als bestehender Stand "
                  f"({existing_stand}) - nichts unternommen.")
            return

        write_alarm(inventar_dir, str(e),
                    f"masterlist.json wurde stattdessen aus der Fallback-Datei "
                    f"{input_path.name} (Stand {fallback_date}) erzeugt.")
        data = convert_old(input_path)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    json_str = json.dumps(data, ensure_ascii=False, separators=(',', ':'))
    out_path.write_text(json_str, encoding='utf-8')

    size_kb = len(json_str.encode()) / 1024
    print(f"OK  Gespeichert: {out_path}  ({size_kb:.0f} KB)  Quelle: {data['_source']}")

    if not no_push:
        is_fallback = data['_source'].startswith('fallback:')
        git_push(out_path, data['_generated'], is_fallback)


def main():
    parser = argparse.ArgumentParser(description='Inventarlisten -> masterlist.json für Inventur-PWA')
    parser.add_argument('--input', default=str(DEFAULT_INPUT),
                        help='Pfad zur alten Fallback-XLSX (nur genutzt wenn wöchentliche Quelle '
                             'fehlt/kaputt ist UND diese Datei neuer als der bestehende masterlist.json-Stand ist)')
    parser.add_argument('--inventar-dir', default=str(DEFAULT_INVENTAR),
                        help='Ordner mit den wöchentlichen Inventarlisten (primäre Quelle)')
    parser.add_argument('--out', default=str(DEFAULT_OUT),
                        help='Ausgabepfad für masterlist.json')
    parser.add_argument('--no-push', action='store_true',
                        help='Kein automatischer Git-Commit + Push')
    parser.add_argument('--watch', action='store_true',
                        help='Quellen überwachen, bei Änderung automatisch konvertieren + pushen')
    parser.add_argument('--interval', type=int, default=30,
                        help='Prüfintervall in Sekunden (nur mit --watch, default: 30)')
    args = parser.parse_args()

    input_path    = Path(args.input)
    inventar_dir  = Path(args.inventar_dir)
    out_path      = Path(args.out)

    if args.watch:
        run_watch(input_path, inventar_dir, out_path, args.no_push, args.interval)
        return

    convert(input_path, inventar_dir, out_path, args.no_push)


def git_push(json_path: Path, date_str: str, is_fallback: bool = False):
    """Commit masterlist.json und push zu GitHub."""
    repo_dir = json_path.parent.parent  # inventur-pwa/

    def run(cmd):
        result = subprocess.run(cmd, cwd=str(repo_dir), capture_output=True, text=True)
        if result.returncode != 0:
            raise RuntimeError(result.stderr.strip() or result.stdout.strip())
        return result.stdout.strip()

    print()
    try:
        # Nur pushen wenn sich die Datei tatsächlich geändert hat
        changed = run(['git', 'diff', '--name-only', str(json_path.relative_to(repo_dir))])
        if not changed:
            # Auch untracked prüfen
            status = run(['git', 'status', '--porcelain', str(json_path.relative_to(repo_dir))])
            if not status:
                print("Git: keine Aenderungen, kein Push noetig.")
                return

        suffix = ' (Fallback: alte Quelle wg. Format-Problem)' if is_fallback else ''
        run(['git', 'add', str(json_path.relative_to(repo_dir))])
        run(['git', 'commit', '-m', f'Masterlist Update {date_str}{suffix}'])
        run(['git', 'push'])
        print("Git: committed + pushed -> GitHub Pages wird aktualisiert.")
    except RuntimeError as e:
        print(f"Git-Fehler: {e}")
        print("Manuell pushen: git add data/masterlist.json && git commit -m 'Update' && git push")


if __name__ == '__main__':
    main()
